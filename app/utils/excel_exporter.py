# -*- coding: utf-8 -*-
"""Unified Excel report generator for Alto Fulfillment.

Uses openpyxl + tempfile for minimal memory footprint.
Error rows (overpacking) are highlighted in red.

Reports:
  - generate_packing_report()           — per-reception packing report
  - generate_reception_report()         — reception scan report
  - generate_employee_packing_report()  — global employee packing journal
"""

import gc
import os
import tempfile
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import (
    Shipment, Item, Packed, PackingLog, ReceptionItem, User,
)

# ──────────────────────────────────────────────
#  Reusable styles (instantiated once per import)
# ──────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_BORDER = Border(bottom=Side(style="thin", color="1E40AF"))

_ERROR_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_ERROR_FONT = Font(color="991B1B")

_TITLE_FONT = Font(bold=True, size=13, color="1E3A5F")

_TOTAL_FONT = Font(bold=True, size=11)
_TOTAL_FILL = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")


# ──────────────────────────────────────────────
#  Helpers
# ──────────────────────────────────────────────

def _fmt_dt(dt: Optional[datetime]) -> str:
    """Format datetime for Excel cell."""
    if not dt:
        return "—"
    return dt.strftime("%d.%m.%Y %H:%M:%S")


def _fmt_duration(seconds: Optional[float]) -> str:
    """Human-readable duration string."""
    if not seconds or seconds <= 0:
        return "—"
    minutes = int(seconds // 60)
    secs = int(seconds % 60)
    if minutes >= 60:
        hours = minutes // 60
        mins = minutes % 60
        return f"{hours}ч {mins}м {secs}с"
    return f"{minutes}м {secs}с"


def _safe_temp_file() -> str:
    """Create a safe temp file path for Excel output."""
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tf.close()
    return tf.name


def _write_header(ws, headers: list[str], widths: list[int], row: int = 2):
    """Write a styled header row."""
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _HEADER_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[row].height = 28


def _write_title(ws, title: str, merge_range: str = "A1:I1"):
    """Write a styled title row."""
    ws.merge_cells(merge_range)
    cell = ws["A1"]
    cell.value = title
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30


# ──────────────────────────────────────────────
#  Report: Packing per reception
# ──────────────────────────────────────────────

def generate_packing_report(db: Session, reception_id: int) -> str:
    """Generate packing report for a single reception.

    Returns: path to temporary .xlsx file (caller must delete after sending).

    Columns: Баркод | Товар | Размер | План | Упаковано (факт) |
             Имя сотрудника | Время начала | Время конца | Длительность
    """
    rec = db.query(Shipment).filter(
        Shipment.id == reception_id, Shipment.type == "reception"
    ).first()
    if not rec:
        raise ValueError(f"Reception #{reception_id} not found")

    logs = (
        db.query(PackingLog, User.full_name, User.username)
        .outerjoin(User, PackingLog.packed_by == User.id)
        .filter(PackingLog.reception_id == reception_id)
        .order_by(PackingLog.created_at)
        .all()
    )

    items_map = {}
    for item in db.query(ReceptionItem).filter(
        ReceptionItem.shipment_id == reception_id
    ).all():
        items_map[item.barcode] = item

    temp_path = _safe_temp_file()

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт упаковки"

        _write_title(ws, f"Отчёт упаковки из приёмки: {rec.name}  (#{rec.id})")

        headers = [
            "Баркод", "Товар", "Размер", "План",
            "Упаковано (факт)", "Имя сотрудника",
            "Время начала", "Время конца", "Длительность",
        ]
        col_widths = [18, 25, 12, 8, 16, 22, 20, 20, 14]
        _write_header(ws, headers, col_widths)

        row_num = 3
        total_plan = 0
        total_fact = 0
        error_count = 0

        for log_entry, full_name, username in logs:
            item = items_map.get(log_entry.barcode)
            article = item.article if item else "—"
            size = item.size if item and item.size and item.size not in ("—", "None") else "—"
            worker_name = full_name or username or "—"

            values = [
                log_entry.barcode, article, size,
                log_entry.plan_quantity, log_entry.quantity, worker_name,
                _fmt_dt(log_entry.start_time), _fmt_dt(log_entry.end_time),
                _fmt_duration(log_entry.duration_seconds),
            ]

            is_error = log_entry.is_error == 1

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                if is_error:
                    cell.fill = _ERROR_FILL
                    cell.font = _ERROR_FONT

            total_plan += log_entry.plan_quantity
            total_fact += log_entry.quantity
            if is_error:
                error_count += 1
            row_num += 1

        # Items not yet packed
        packed_barcodes = {entry.barcode for entry, _, _ in logs}
        for barcode, item in items_map.items():
            if barcode not in packed_barcodes:
                article = item.article or "—"
                size = item.size if item.size and item.size not in ("—", "None") else "—"
                values = [barcode, article, size, item.quantity, 0, "—", "—", "—", "—"]
                for col_idx, val in enumerate(values, 1):
                    ws.cell(row=row_num, column=col_idx, value=val)
                row_num += 1
                total_plan += item.quantity

        # Totals
        row_num += 1
        total_values = ["ИТОГО", "", "", total_plan, total_fact, "", "", "", ""]
        for col_idx, val in enumerate(total_values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = _TOTAL_FONT
            cell.fill = _TOTAL_FILL

        if error_count > 0:
            row_num += 1
            err_cell = ws.cell(row=row_num, column=1, value=f"⚠ Ошибок (перебор): {error_count}")
            err_cell.font = Font(bold=True, color="DC2626")

        wb.save(temp_path)
        wb.close()

    except Exception:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise
    finally:
        gc.collect()

    return temp_path


# ──────────────────────────────────────────────
#  Report: Reception scan report
# ──────────────────────────────────────────────

def generate_reception_report(db: Session, reception_id: int) -> str:
    """Generate reception scan report.

    Returns: path to temporary .xlsx file.
    """
    rec = db.query(Shipment).filter(
        Shipment.id == reception_id, Shipment.type == "reception"
    ).first()
    if not rec:
        raise ValueError(f"Reception #{reception_id} not found")

    temp_path = _safe_temp_file()

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Приёмка"

        _write_title(ws, f"Приёмка: {rec.name}  (#{rec.id})", merge_range="A1:F1")

        headers = ["Баркод", "Товар", "Размер", "План", "Факт", "Номер короба"]
        col_widths = [18, 25, 12, 10, 10, 14]
        _write_header(ws, headers, col_widths)

        query = (
            db.query(ReceptionItem)
            .filter(ReceptionItem.shipment_id == reception_id)
            .order_by(ReceptionItem.barcode)
        )

        row_num = 3
        sum_plan = 0
        sum_fact = 0

        for it in query.yield_per(100):
            barcode = it.barcode
            article = it.article or "—"
            size = it.size if it.size and it.size not in ("—", "None") else "—"
            plan = it.quantity
            fact = it.quantity
            box = getattr(it, "box_number", "")

            for col_idx, val in enumerate([barcode, article, size, plan, fact, box], 1):
                ws.cell(row=row_num, column=col_idx, value=val)

            sum_plan += plan
            sum_fact += fact
            row_num += 1

        # Totals
        row_num += 1
        total_values = ["ИТОГО", "", "", sum_plan, sum_fact, ""]
        for col_idx, val in enumerate(total_values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = _TOTAL_FONT
            cell.fill = _TOTAL_FILL

        wb.save(temp_path)
        wb.close()

    except Exception:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise
    finally:
        gc.collect()

    return temp_path


# ──────────────────────────────────────────────
#  Report: Employee Packing Journal (global)
# ──────────────────────────────────────────────

def generate_employee_packing_report(db: Session, logs_query) -> str:
    """Generate employee packing journal across all receptions.

    Columns: ФИО | Дата и Время | Баркод | Упаковано (факт) | План |
             Разница | Время работы (сек) | ТЗ (Приёмка) | ID Приёмки

    Rows where fact > plan are highlighted red (overpacking).
    Returns: path to temporary .xlsx file.
    """
    temp_path = _safe_temp_file()

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Журнал упаковки"

        _write_title(
            ws,
            f"Журнал упаковки сотрудников — {datetime.utcnow().strftime('%d.%m.%Y %H:%M')}",
            merge_range="A1:I1",
        )

        headers = [
            "ФИО сотрудника",
            "Дата и Время",
            "Баркод",
            "Упаковано (факт)",
            "План",
            "Разница",
            "Время работы (сек)",
            "ТЗ (Приёмка)",
            "ID Приёмки",
        ]
        col_widths = [24, 20, 18, 16, 10, 12, 18, 25, 12]
        _write_header(ws, headers, col_widths)

        row_num = 3
        total_fact = 0
        total_plan = 0
        error_count = 0

        for row in logs_query.yield_per(500):
            # Unpack query columns
            (
                log_id, full_name, username, tz_name,
                start_time, end_time, duration_seconds,
                barcode, quantity, plan_quantity,
            ) = row

            worker = full_name or username or "—"
            diff = plan_quantity - quantity
            is_overpack = quantity > plan_quantity

            values = [
                worker,
                _fmt_dt(end_time or start_time),
                barcode,
                quantity,
                plan_quantity,
                diff,
                round(duration_seconds, 1) if duration_seconds else "—",
                tz_name or "—",
                log_id,
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                if is_overpack:
                    cell.fill = _ERROR_FILL
                    cell.font = _ERROR_FONT

            total_fact += quantity
            total_plan += plan_quantity
            if is_overpack:
                error_count += 1
            row_num += 1

        # Totals
        row_num += 1
        total_values = [
            "ИТОГО", "", "",
            total_fact, total_plan,
            total_fact - total_plan,
            "", "", "",
        ]
        for col_idx, val in enumerate(total_values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = _TOTAL_FONT
            cell.fill = _TOTAL_FILL

        if error_count > 0:
            row_num += 1
            err_cell = ws.cell(
                row=row_num, column=1,
                value=f"⚠ Случаев перебора: {error_count}",
            )
            err_cell.font = Font(bold=True, color="DC2626")

        wb.save(temp_path)
        wb.close()

    except Exception:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise
    finally:
        gc.collect()

    return temp_path
