# -*- coding: utf-8 -*-
"""Stock router: warehouse inventory management."""

import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import User, Stock
from app.auth import get_current_user

router = APIRouter(prefix="/api/stock", tags=["stock"])


@router.get("")
def list_stock(
    show_archived: bool = False,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List suppliers with stock totals."""
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    is_archived = 1 if show_archived else 0
    suppliers = (
        db.query(
            Stock.supplier_name,
            func.sum(Stock.quantity).label("total"),
            func.count(Stock.id).label("positions"),
            func.max(Stock.updated_at).label("last_upd"),
        )
        .filter(Stock.is_archived == is_archived)
        .group_by(Stock.supplier_name)
        .order_by(Stock.supplier_name)
        .all()
    )
    return [
        {
            "supplier_name": name,
            "total": total,
            "positions": positions,
            "updated_at": last_upd.strftime("%d.%m.%Y %H:%M") if last_upd else None,
        }
        for name, total, positions, last_upd in suppliers
    ]


@router.get("/{supplier_name}/report")
def stock_report(
    supplier_name: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Download Excel report for supplier stock."""
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    import openpyxl

    items = (
        db.query(Stock)
        .filter(Stock.supplier_name == supplier_name)
        .order_by(Stock.barcode)
        .all()
    )
    if not items:
        raise HTTPException(status_code=404, detail=f"У поставщика «{supplier_name}» нет записей")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Остатки — {supplier_name}"

    header_fill = openpyxl.styles.PatternFill(start_color="D9E2F3", fill_type="solid")
    bold_font = openpyxl.styles.Font(bold=True)

    headers = ["Штрихкод", "Артикул", "Размер", "Остаток"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.fill = header_fill

    for idx, s in enumerate(items, 2):
        ws.cell(row=idx, column=1, value=s.barcode)
        ws.cell(row=idx, column=2, value=s.article or "—")
        ws.cell(row=idx, column=3, value=s.size or "—")
        ws.cell(row=idx, column=4, value=s.quantity)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"stock_{supplier_name}_{datetime.now().strftime('%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{supplier_name}/archive")
def archive_supplier(
    supplier_name: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Archive supplier stock entries."""
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    updated = (
        db.query(Stock)
        .filter(Stock.supplier_name == supplier_name, Stock.is_archived == 0)
        .update({Stock.is_archived: 1})
    )
    db.commit()
    if not updated:
        raise HTTPException(status_code=404, detail="Поставщик не найден")
    return {"status": "ok", "message": f"Поставщик «{supplier_name}» перемещён в архив"}


@router.post("/{supplier_name}/unarchive")
def unarchive_supplier(
    supplier_name: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Unarchive supplier stock entries."""
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    updated = (
        db.query(Stock)
        .filter(Stock.supplier_name == supplier_name, Stock.is_archived == 1)
        .update({Stock.is_archived: 0})
    )
    db.commit()
    if not updated:
        raise HTTPException(status_code=404, detail="Поставщик не найден")
    return {"status": "ok", "message": f"Поставщик «{supplier_name}» возвращён из архива"}


@router.delete("/{supplier_name}")
def delete_supplier(
    supplier_name: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Delete all stock entries for supplier. Admin and Manager."""
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    deleted = db.query(Stock).filter(Stock.supplier_name == supplier_name).delete()
    db.commit()
    if not deleted:
        raise HTTPException(status_code=404, detail="Поставщик не найден")
    return {"status": "ok", "message": f"Поставщик «{supplier_name}» удалён"}
