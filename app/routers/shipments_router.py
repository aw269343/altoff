# -*- coding: utf-8 -*-
"""Shipments router: create from Excel, scan, pack with KPI, box management, archive, export report."""

import io
import os
import gc
from datetime import datetime
from typing import Optional
from urllib.parse import quote as urllib_quote

import openpyxl

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy import func, distinct
from sqlalchemy.orm import Session
from starlette.background import BackgroundTask

from app.database import get_db
from app.models import User, Shipment, Item, Packed, Stock, StockMovement, PackingLog
from app.auth import get_current_user
from app.utils.excel_exporter import generate_packing_report

router = APIRouter(prefix="/api/shipments", tags=["shipments"])


# ---------- Schemas ----------

class ShipmentListItem(BaseModel):
    id: int
    name: str
    status: str
    last_box_number: int
    created_at: Optional[str] = None
    total_plan: int = 0
    total_packed: int = 0
    box_count: int = 0


class ItemDetail(BaseModel):
    id: int
    barcode: str
    quantity: int
    packed: int
    remaining: int
    article: str
    size: str


class ShipmentDetail(BaseModel):
    id: int
    name: str
    status: str
    last_box_number: int
    items: list


class ScanRequest(BaseModel):
    box_number: int
    item_barcode: str
    quantity: int = 1


class PackRequest(BaseModel):
    """Pack request with timing data for KPI tracking."""
    barcode: str
    quantity: int
    box_number: int
    start_time: Optional[str] = None  # ISO datetime string from frontend


# ---------- Endpoints ----------

@router.get("")
def list_shipments(
    status_filter: str = "active,completed",
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List shipments with given statuses."""
    statuses = [s.strip() for s in status_filter.split(",")]
    shipments = (
        db.query(Shipment)
        .filter(Shipment.status.in_(statuses), Shipment.type == "shipment")
        .order_by(Shipment.created_at.desc())
        .all()
    )
    result = []
    for s in shipments:
        total_plan = db.query(func.sum(Item.quantity)).filter(Item.shipment_id == s.id).scalar() or 0
        total_packed = db.query(func.count(Packed.id)).filter(Packed.shipment_id == s.id).scalar() or 0
        box_count = db.query(func.count(distinct(Packed.box_number))).filter(Packed.shipment_id == s.id).scalar() or 0
        result.append(ShipmentListItem(
            id=s.id,
            name=s.name,
            status=s.status,
            last_box_number=s.last_box_number,
            created_at=s.created_at.strftime("%d.%m.%Y %H:%M") if s.created_at else None,
            total_plan=total_plan,
            total_packed=total_packed,
            box_count=box_count,
        ))
    return result


@router.get("/{shipment_id}")
def get_shipment(
    shipment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get shipment detail with plan + progress."""
    shipment = db.query(Shipment).filter(Shipment.id == shipment_id, Shipment.type == "shipment").first()
    if not shipment:
        raise HTTPException(status_code=404, detail="Поставка не найдена")

    total_plan = 0
    total_packed = 0
    items_data = []
    for item in shipment.items:
        packed_count = (
            db.query(func.count(Packed.id))
            .filter(Packed.shipment_id == shipment_id, Packed.barcode == item.barcode)
            .scalar()
        )
        total_plan += item.quantity
        total_packed += packed_count
        items_data.append({
            "id": item.id,
            "barcode": item.barcode,
            "quantity": item.quantity,
            "packed": packed_count,
            "remaining": item.quantity - packed_count,
            "article": item.article or "—",
            "size": item.size or "—",
        })

    return {
        "id": shipment.id,
        "name": shipment.name,
        "status": shipment.status,
        "last_box_number": shipment.last_box_number,
        "total_plan": total_plan,
        "total_packed": total_packed,
        "items": items_data,
    }


@router.post("/upload")
def upload_shipment(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Create shipment from Excel upload. Duplicate barcodes are summed."""
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Нужен файл формата .xlsx")

    data = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        items = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if len(row) < 2:
                continue
            barcode_val, qty_val = row[0], row[1]
            if barcode_val is None or qty_val is None:
                continue
            barcode = str(int(barcode_val)) if isinstance(barcode_val, float) else str(barcode_val)
            barcode = barcode.strip()
            try:
                qty = int(qty_val)
            except (ValueError, TypeError):
                continue
            article_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
            size_val = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
            if qty > 0 and barcode:
                items.append((barcode, qty, article_val, size_val))
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка чтения Excel: {e}")

    if not items:
        raise HTTPException(status_code=400, detail="В файле не найдено ни одного товара")

    # Group duplicates — sum quantities
    grouped = {}
    for barcode, qty, article_val, size_val in items:
        if barcode in grouped:
            grouped[barcode] = (grouped[barcode][0] + qty, grouped[barcode][1], grouped[barcode][2])
        else:
            grouped[barcode] = (qty, article_val, size_val)
    items = [(bc, info[0], info[1], info[2]) for bc, info in grouped.items()]

    shipment_name = os.path.splitext(file.filename)[0]

    shipment = Shipment(
        name=shipment_name,
        type="shipment",
        status="active",
        created_by=current_user.id,
    )
    db.add(shipment)
    db.flush()

    for barcode, qty, article, size in items:
        db.add(Item(
            shipment_id=shipment.id,
            barcode=barcode,
            quantity=qty,
            article=article,
            size=size,
        ))

    db.commit()
    total_items = sum(q for _, q, _, _ in items)
    return {
        "status": "ok",
        "id": shipment.id,
        "name": shipment_name,
        "articles": len(items),
        "total": total_items,
    }


@router.post("/{shipment_id}/scan")
def scan_item(
    shipment_id: int,
    body: ScanRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Scan an item into a box (legacy endpoint, kept for compatibility)."""
    quantity = max(1, body.quantity)

    item = db.query(Item).filter(
        Item.shipment_id == shipment_id,
        Item.barcode == body.item_barcode,
    ).first()

    if not item:
        raise HTTPException(status_code=404, detail=f"Артикул {body.item_barcode} не найден в плане")

    packed_count = (
        db.query(func.count(Packed.id))
        .filter(Packed.shipment_id == shipment_id, Packed.barcode == body.item_barcode)
        .scalar()
    )

    if packed_count + quantity > item.quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Перебор! {body.item_barcode}: план {item.quantity}, уже {packed_count}, добавляете {quantity}",
        )

    for _ in range(quantity):
        db.add(Packed(
            shipment_id=shipment_id,
            item_id=item.id,
            barcode=body.item_barcode,
            box_number=body.box_number,
            packed_by=current_user.id,
        ))

    db.commit()
    remaining = item.quantity - packed_count - quantity
    return {"status": "ok", "remaining": remaining}




@router.post("/{shipment_id}/open-box")
def open_box(
    shipment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get current box number (create first if 0)."""
    shipment = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="Поставка не найдена")

    if shipment.last_box_number == 0:
        shipment.last_box_number = 1
        db.commit()

    return {"status": "ok", "box_number": shipment.last_box_number}


@router.post("/{shipment_id}/close-box")
def close_box(
    shipment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Close current box → next number. Auto-complete if all items packed."""
    shipment = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="Поставка не найдена")

    shipment.last_box_number += 1
    new_box = shipment.last_box_number

    # Check all items complete
    all_complete = True
    for item in shipment.items:
        packed_count = (
            db.query(func.count(Packed.id))
            .filter(Packed.shipment_id == shipment_id, Packed.barcode == item.barcode)
            .scalar()
        )
        if packed_count < item.quantity:
            all_complete = False
            break

    if all_complete and len(shipment.items) > 0:
        shipment.status = "completed"

    db.commit()
    return {"status": "ok", "box_number": new_box, "completed": all_complete}


@router.post("/{shipment_id}/add-items")
def add_items_to_shipment(
    shipment_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Add additional items to an existing shipment from Excel."""
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    shipment = db.query(Shipment).filter(
        Shipment.id == shipment_id, Shipment.type == "shipment"
    ).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="Поставка не найдена")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Нужен файл формата .xlsx")

    data = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        new_items = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if len(row) < 2:
                continue
            barcode_val, qty_val = row[0], row[1]
            if barcode_val is None or qty_val is None:
                continue
            barcode = str(int(barcode_val)) if isinstance(barcode_val, float) else str(barcode_val)
            barcode = barcode.strip()
            try:
                qty = int(qty_val)
            except (ValueError, TypeError):
                continue
            article_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
            size_val = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
            if qty > 0 and barcode:
                new_items.append((barcode, qty, article_val, size_val))
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка чтения Excel: {e}")

    if not new_items:
        raise HTTPException(status_code=400, detail="В файле не найдено ни одного товара")

    # Group duplicates within the uploaded file
    grouped = {}
    for barcode, qty, article_val, size_val in new_items:
        if barcode in grouped:
            grouped[barcode] = (grouped[barcode][0] + qty, grouped[barcode][1], grouped[barcode][2])
        else:
            grouped[barcode] = (qty, article_val, size_val)

    # Add to existing shipment (merge with existing items or create new)
    added_count = 0
    added_qty = 0
    for barcode, (qty, article, size) in grouped.items():
        existing = db.query(Item).filter(
            Item.shipment_id == shipment_id, Item.barcode == barcode
        ).first()
        if existing:
            existing.quantity += qty
            if article and (not existing.article or existing.article == "—"):
                existing.article = article
            if size and (not existing.size or existing.size == "—"):
                existing.size = size
        else:
            db.add(Item(
                shipment_id=shipment_id,
                barcode=barcode,
                quantity=qty,
                article=article,
                size=size,
            ))
            added_count += 1
        added_qty += qty

    db.commit()
    return {
        "status": "ok",
        "new_articles": added_count,
        "added_qty": added_qty,
        "total_articles": len(grouped),
    }




@router.delete("/{shipment_id}")
def delete_shipment(
    shipment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Delete shipment. Admin and Manager."""
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    shipment = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="Поставка не найдена")

    db.delete(shipment)
    db.commit()
    return {"status": "ok", "message": f"Поставка #{shipment_id} удалена"}
