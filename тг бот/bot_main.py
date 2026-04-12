#!/opt/python/python-3.10.1/bin/python
# -*- coding: utf-8 -*-
# main.py — Telegram-бот "Alto" (фулфилмент) + Mini App API
#
# Режим работы: CGI (Reg.ru / ISPmanager)
# Скрипт запускается заново при КАЖДОМ HTTP-запросе.
#
# Роли: Admin → Manager → Storekeeper
# Управление пользователями по @username
# Создание поставок через загрузку Excel
# Автонумерация коробов
#
# Маршрутизация:
#   Использует режим Long Polling (getUpdates).
#
# Роли: Admin → Manager → Storekeeper
# Управление пользователями по @username
# Создание поставок через загрузку Excel
# Автонумерация коробов
#
# Зависимости: pip install sqlalchemy openpyxl

import sys
import os
import io
import json
import urllib.request
import urllib.parse
import tempfile
from datetime import datetime

# ---------------------------------------------------------------------------
#  Абсолютные пути
# ---------------------------------------------------------------------------
SCRIPT_DIR = "/var/www/u3262373/data/www/altoff.online"
sys.path.insert(0, SCRIPT_DIR)

from database import init_db, get_session, User, Shipment, Item, Packed, ReceptionItem, Stock, StockMovement
from sqlalchemy import func, distinct

# ---------------------------------------------------------------------------
#  Конфигурация
# ---------------------------------------------------------------------------

BOT_TOKEN = "8540310746:AAFvN8vkajepRpAjQ-K_zJe-cfxmh4ZNghk"
ADMIN_ID = 397821847
ADMIN_USERNAMES = ["iambelaya"]
ADMIN_USERNAME = ""   # заполнится при первом /start от админа
WEBHOOK_HOST = "https://altoff.online"
MINI_APP_URL = "https://altoff.online/login"

STATE_FILE = os.path.join(SCRIPT_DIR, "user_states.json")
LOG_FILE = os.path.join(SCRIPT_DIR, "log.txt")

# ---------------------------------------------------------------------------
#  Инициализация БД
# ---------------------------------------------------------------------------
init_db()


# ===================================================================
#  CGI-утилиты
# ===================================================================

def get_request_method():
    return os.environ.get("REQUEST_METHOD", "GET")


def get_query_params():
    qs = os.environ.get("QUERY_STRING", "")
    return urllib.parse.parse_qs(qs)


def read_request_body():
    content_length = int(os.environ.get("CONTENT_LENGTH", 0))
    if content_length > 0:
        return sys.stdin.buffer.read(content_length)
    return b""


def respond_json(data, status="200 OK"):
    body = json.dumps(data, ensure_ascii=False)
    sys.stdout.buffer.write(f"Status: {status}\r\n".encode("utf-8"))
    sys.stdout.buffer.write(b"Content-Type: application/json; charset=utf-8\r\n")
    sys.stdout.buffer.write(f"Content-Length: {len(body.encode('utf-8'))}\r\n".encode("utf-8"))
    sys.stdout.buffer.write(b"\r\n")
    sys.stdout.buffer.write(body.encode("utf-8"))
    sys.stdout.buffer.flush()


def respond_html(html_content, status="200 OK"):
    body = html_content.encode("utf-8")
    sys.stdout.buffer.write(f"Status: {status}\r\n".encode("utf-8"))
    sys.stdout.buffer.write(b"Content-Type: text/html; charset=utf-8\r\n")
    sys.stdout.buffer.write(f"Content-Length: {len(body)}\r\n".encode("utf-8"))
    sys.stdout.buffer.write(b"\r\n")
    sys.stdout.buffer.write(body)
    sys.stdout.buffer.flush()


# ===================================================================
#  Логирование
# ===================================================================

def log(text):
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now()}] {text}\n")
    except Exception:
        pass


# ===================================================================
#  Telegram Bot API
# ===================================================================

def tg_api(method, payload=None):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/{method}"
    data = json.dumps(payload or {}).encode("utf-8")
    req = urllib.request.Request(
        url, data=data,
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        log(f"TG API error ({method}): {e}")
        return None


def send_message(chat_id, text, parse_mode="HTML", reply_markup=None):
    payload = {"chat_id": chat_id, "text": text, "parse_mode": parse_mode}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    return tg_api("sendMessage", payload)


def send_document(chat_id, file_bytes, filename, caption=""):
    boundary = "----FormBoundary7MA4YWxkTrZu0gW"
    body = b""
    body += f"--{boundary}\r\n".encode()
    body += b"Content-Disposition: form-data; name=\"chat_id\"\r\n\r\n"
    body += f"{chat_id}\r\n".encode()
    if caption:
        body += f"--{boundary}\r\n".encode()
        body += b"Content-Disposition: form-data; name=\"caption\"\r\n\r\n"
        body += f"{caption}\r\n".encode()
    body += f"--{boundary}\r\n".encode()
    body += b"Content-Disposition: form-data; name=\"parse_mode\"\r\n\r\n"
    body += b"HTML\r\n"
    body += f"--{boundary}\r\n".encode()
    body += f"Content-Disposition: form-data; name=\"document\"; filename=\"{filename}\"\r\n".encode()
    body += b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    body += file_bytes
    body += b"\r\n"
    body += f"--{boundary}--\r\n".encode()

    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
    req = urllib.request.Request(url, data=body, headers={
        "Content-Type": f"multipart/form-data; boundary={boundary}",
    })
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        log(f"TG sendDocument error: {e}")
        return None


def answer_callback(callback_query_id, text=""):
    tg_api("answerCallbackQuery", {"callback_query_id": callback_query_id, "text": text})


def edit_message(chat_id, message_id, text, reply_markup=None):
    payload = {"chat_id": chat_id, "message_id": message_id,
               "text": text, "parse_mode": "HTML"}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    return tg_api("editMessageText", payload)


def delete_message(chat_id, message_id):
    tg_api("deleteMessage", {"chat_id": chat_id, "message_id": message_id})


def download_tg_file(file_id):
    """Скачать файл с серверов Telegram."""
    result = tg_api("getFile", {"file_id": file_id})
    if not result or not result.get("ok"):
        return None
    file_path = result["result"]["file_path"]
    url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}"
    try:
        with urllib.request.urlopen(url, timeout=30) as resp:
            return resp.read()
    except Exception as e:
        log(f"Download file error: {e}")
        return None


# ===================================================================
#  FSM (состояния через файл)
# ===================================================================

def load_states():
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_states(states):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(states, f, ensure_ascii=False)


def get_user_state(user_id):
    return load_states().get(str(user_id))


def set_user_state(user_id, state):
    states = load_states()
    states[str(user_id)] = state
    save_states(states)


def clear_user_state(user_id):
    states = load_states()
    states.pop(str(user_id), None)
    save_states(states)


# ===================================================================
#  Роли и клавиатуры
# ===================================================================

def get_user_role(user_id, username=None):
    """Определить роль пользователя."""
    if user_id == ADMIN_ID:
        return "admin"
    if username and username.lower() in ADMIN_USERNAMES:
        return "admin"
    with get_session() as session:
        user = session.query(User).filter(User.telegram_id == user_id).first()
        if user:
            return user.role
        # Попробовать найти по username (до первого /start telegram_id может быть неизвестен)
        if username:
            user = session.query(User).filter(User.username == username.lower()).first()
            if user:
                # Обновить telegram_id при первом обращении
                user.telegram_id = user_id
                return user.role
    return None


def keyboard_for_role(role):
    """Клавиатура с наследованием прав: Admin > Manager > Storekeeper."""
    rows = []

    # --- Ряд 1: Создание (Manager + Admin, ожидание Excel) ---
    if role in ("admin", "manager"):
        rows.append([
            {"text": "📥 Создать приёмку"},
            {"text": "📦 Создать поставку"},
        ])

    # --- Ряд 2: Работа (все роли, Mini App) ---
    rows.append([
        {"text": "📲 Открыть склад (Приёмка)", "web_app": {"url": "https://altoff.online/login"}},
        {"text": "🛠️ Открыть склад (Сборка)", "web_app": {"url": "https://altoff.online/login"}},
    ])

    # --- Ряд 3-4: Менеджер + Админ ---
    if role in ("admin", "manager"):
        rows.append([{"text": "� Учёт товара"}, {"text": "📋 Мои поставки"}])
        rows.append([{"text": "� История приёмок"}, {"text": "📜 История поставок"}])

    # --- Ряд 5-6: Только Админ (Менеджеры, Удаление) ---
    if role == "admin":
        rows.append([{"text": "👥 Менеджеры"}, {"text": "❌ Удалить пользователя"}])

    # --- Ряд 7: Менеджер + Админ (Кладовщики) ---
    if role in ("admin", "manager"):
        rows.append([{"text": "👤 Кладовщики"}])

    return {"keyboard": rows, "resize_keyboard": True}


def cancel_keyboard():
    """Клавиатура с кнопкой «Отмена» для FSM-шагов."""
    return {
        "keyboard": [[{"text": "❌ Отмена"}]],
        "resize_keyboard": True,
    }


# ===================================================================
#  Обработка Telegram-обновлений
# ===================================================================

def handle_telegram_update(update):
    log(f"Update: {json.dumps(update, ensure_ascii=False)}")

    # --- Callback Query (inline-кнопки) ---
    callback = update.get("callback_query")
    if callback:
        handle_callback(callback)
        return

    message = update.get("message")
    if not message:
        return

    chat_id = message["chat"]["id"]
    user_id = message["from"]["id"]
    username = (message["from"].get("username") or "").lower()
    text = (message.get("text") or "").strip()

    # --- Определить роль ---
    role = get_user_role(user_id, username)
    if not role:
        send_message(chat_id, "🚫 Доступ запрещён. Обратитесь к администратору.")
        return

    # --- Кнопка «Отмена» — выход из любого шага ---
    state = get_user_state(user_id)
    if text == "❌ Отмена":
        if state:
            clear_user_state(user_id)
        send_message(chat_id, "↩️ Действие отменено. Главное меню:",
                     reply_markup=keyboard_for_role(role))
        return

    # --- FSM: ожидание документа (создание поставки) ---
    if state == "waiting_excel":
        doc = message.get("document")
        if doc:
            process_excel_upload(chat_id, user_id, doc, role)
        else:
            send_message(chat_id, "📎 Пожалуйста, отправьте Excel-файл (.xlsx).",
                         reply_markup=cancel_keyboard())
        return

    # --- FSM: ожидание документа (импорт приёмки) ---
    if state == "waiting_reception_excel":
        doc = message.get("document")
        if doc:
            process_reception_excel(chat_id, user_id, doc, role)
        else:
            send_message(chat_id, "📎 Пожалуйста, отправьте Excel-файл (.xlsx) для приёмки.",
                         reply_markup=cancel_keyboard())
        return

    if state and state.startswith("waiting_username_"):
        process_username_input(chat_id, user_id, text, state, role)
        return

    # --- Команды ---
    if text.startswith("/start"):
        welcome = (
            "👋 Добро пожаловать в систему фулфилмента <b>Alto</b>!\n\n"
            f"Ваша роль: <b>{role.upper()}</b>"
        )
        send_message(chat_id, welcome, reply_markup=keyboard_for_role(role))
        return

    # --- Кнопки Admin (Менеджеры, Удаление) ---
    if role == "admin":
        if text == "👥 Менеджеры":
            send_message(chat_id, "Введите @username нового менеджера:",
                         reply_markup=cancel_keyboard())
            set_user_state(user_id, "waiting_username_add_manager")
            return
        if text == "❌ Удалить пользователя":
            send_message(chat_id, "Введите @username пользователя для удаления:",
                         reply_markup=cancel_keyboard())
            set_user_state(user_id, "waiting_username_delete_user")
            return

    # --- Кнопки Менеджер + Админ (Кладовщики) ---
    if role in ("admin", "manager"):
        if text == "👤 Кладовщики":
            send_message(chat_id, "Введите @username нового кладовщика:",
                         reply_markup=cancel_keyboard())
            set_user_state(user_id, "waiting_username_add_storekeeper")
            return

    # --- Кнопки Создания (Manager + Admin) ---
    if role in ("admin", "manager"):
        if text == "📥 Создать приёмку":
            send_message(chat_id,
                         "📎 Отправьте Excel-файл (.xlsx) для приёмки.\n\n"
                         "• <b>Имя файла</b> = название приёмки\n"
                         "• <b>1-й столбец</b> — штрихкод\n"
                         "• <b>2-й столбец</b> — количество\n"
                         "• <b>3-й столбец</b> — артикул (необяз.)\n"
                         "• <b>4-й столбец</b> — размер (необяз.)",
                         reply_markup=cancel_keyboard())
            set_user_state(user_id, "waiting_reception_excel")
            return
        if text == "📦 Создать поставку":
            send_message(chat_id,
                         "📎 Отправьте Excel-файл (.xlsx) для поставки.\n\n"
                         "• <b>Имя файла</b> = название поставки\n"
                         "• <b>1-й столбец</b> — штрихкод\n"
                         "• <b>2-й столбец</b> — количество\n"
                         "• <b>3-й столбец</b> — артикул (необяз.)\n"
                         "• <b>4-й столбец</b> — размер (необяз.)",
                         reply_markup=cancel_keyboard())
            set_user_state(user_id, "waiting_excel")
            return

    # --- Кнопки Списков и Учёта (Manager + Admin) ---
    if role in ("admin", "manager"):
        if text == "📋 Мои поставки":
            handle_my_shipments(chat_id, caller_role=role)
            return
        if text == "� Учёт товара":
            handle_stock(chat_id, caller_role=role)
            return
        if text == "� История приёмок":
            handle_history_receptions(chat_id, caller_role=role)
            return
        if text == "📜 История поставок":
            handle_history_shipments(chat_id, caller_role=role)
            return

    # --- Автодетект Excel-файла (если отправлен без FSM-состояния) ---
    doc = message.get("document")
    if doc and (doc.get("file_name") or "").lower().endswith(".xlsx"):
        if role in ("admin", "manager"):
            # Менеджер/Админ: спросить тип импорта
            buttons = [[
                {"text": "📦 Поставка (сборка)", "callback_data": "excel_as_shipment"},
                {"text": "📥 Приёмка (склад)", "callback_data": "excel_as_reception"},
            ]]
            send_message(chat_id,
                         "📎 Получен Excel-файл. Что создать?\n\n"
                         "• <b>Поставка</b> — план сборки для кладовщиков\n"
                         "• <b>Приёмка</b> — импорт на склад (Stock)",
                         reply_markup={"inline_keyboard": buttons})
            # Сохраняем file_id во временное состояние
            set_user_state(user_id, f"excel_pending_{doc['file_id']}|{doc.get('file_name', 'import.xlsx')}")
            return
        elif role == "storekeeper":
            # Кладовщик: автоматически как приёмку
            process_reception_excel(chat_id, user_id, doc, role)
            return

    # --- Неизвестное сообщение ---
    send_message(chat_id, "Используйте кнопки 👇", reply_markup=keyboard_for_role(role))


# ===================================================================
#  Callback Query (inline-кнопки для управления поставками)
# ===================================================================

def handle_callback(callback):
    data = callback.get("data", "")
    chat_id = callback["message"]["chat"]["id"]
    message_id = callback["message"]["message_id"]
    user_id = callback["from"]["id"]
    username = (callback["from"].get("username") or "").lower()

    role = get_user_role(user_id, username)
    if not role:
        answer_callback(callback["id"], "⛔ Нет доступа")
        return

    # excel_as_shipment / excel_as_reception — выбор типа для загруженного Excel
    if data == "history_shipments":
        handle_history_shipments(chat_id, role)
        return
    if data == "history_receptions":
        handle_history_receptions(chat_id, role)
        return

    if data in ("excel_as_shipment", "excel_as_reception"):
        state = get_user_state(user_id)
        if not state or not state.startswith("excel_pending_"):
            answer_callback(callback["id"], "❌ Файл устарел, отправьте заново")
            return
        pending = state[len("excel_pending_"):]
        parts = pending.split("|", 1)
        file_id = parts[0]
        file_name = parts[1] if len(parts) > 1 else "import.xlsx"
        clear_user_state(user_id)
        delete_message(chat_id, message_id)
        doc = {"file_id": file_id, "file_name": file_name}
        if data == "excel_as_shipment":
            process_excel_upload(chat_id, user_id, doc, role)
        else:
            process_reception_excel(chat_id, user_id, doc, role)
        answer_callback(callback["id"], "✅ Обработано")
        return

    if role == "storekeeper":
        answer_callback(callback["id"], "⛔ Нет доступа")
        return

    # archive_shipment_<id>  — completed → archived
    if data.startswith("archive_shipment_"):
        sid = int(data.split("_")[-1])
        archive_shipment(chat_id, sid)
        answer_callback(callback["id"], "📂 Перемещено в историю")
        delete_message(chat_id, message_id)
        return

    # archive_reception_<id>  — completed → archived
    if data.startswith("archive_reception_"):
        sid = int(data.split("_")[-1])
        archive_shipment(chat_id, sid)
        answer_callback(callback["id"], "📂 Перемещено в историю")
        delete_message(chat_id, message_id)
        return

    # delete_shipment_<id>  — только Admin
    if data.startswith("delete_shipment_") or data.startswith("delete_reception_"):
        if role != "admin":
            answer_callback(callback["id"], "⛔ Только Администратор")
            return
        sid = int(data.split("_")[-1])
        delete_shipment(chat_id, sid)
        answer_callback(callback["id"], "🗑 Удалено")
        delete_message(chat_id, message_id)
        return

    # report_shipment_<id> — отчёт по поставке
    if data.startswith("report_shipment_"):
        sid = int(data.split("_")[-1])
        send_shipment_report(chat_id, sid)
        answer_callback(callback["id"], "📊 Отчёт отправлен")
        return

    # report_reception_<id> — отчёт по приёмке
    if data.startswith("report_reception_"):
        sid = int(data.split("_")[-1])
        send_reception_report(chat_id, sid)
        answer_callback(callback["id"], "📊 Отчёт отправлен")
        return

    # report_<id>  — универсальный (определяем тип)
    if data.startswith("report_"):
        sid = int(data.split("_")[-1])
        with get_session() as session:
            shipment = session.query(Shipment).filter(Shipment.id == sid).first()
            stype = shipment.type if shipment else "shipment"
        if stype == "reception":
            send_reception_report(chat_id, sid)
        else:
            send_shipment_report(chat_id, sid)
        answer_callback(callback["id"], "📊 Отчёт отправлен")
        return

    # reopen_<id>  — archived → completed (только Admin)
    if data.startswith("reopen_"):
        if role != "admin":
            answer_callback(callback["id"], "⛔ Только Администратор")
            return
        sid = int(data.split("_")[-1])
        reopen_shipment(chat_id, sid)
        answer_callback(callback["id"], "✅ Возвращено")
        delete_message(chat_id, message_id)
        return

    # stock_report_<supplier> — отчёт по остаткам поставщика
    if data.startswith("stock_report_"):
        supplier = data[len("stock_report_"):]
        send_stock_report(chat_id, supplier)
        answer_callback(callback["id"], "📊 Отчёт отправлен")
        return

    # stock_archive_<supplier> — скрыть поставщика из учёта
    if data.startswith("stock_archive_"):
        supplier = data[len("stock_archive_"):]
        archive_stock_supplier(chat_id, supplier)
        answer_callback(callback["id"], "📂 В архив")
        delete_message(chat_id, message_id)
        return

    # stock_unarchive_<supplier> — вернуть поставщика из архива
    if data.startswith("stock_unarchive_"):
        supplier = data[len("stock_unarchive_"):]
        unarchive_stock_supplier(chat_id, supplier)
        answer_callback(callback["id"], "🔄 Возвращено")
        delete_message(chat_id, message_id)
        return

    # stock_delete_<supplier> — удалить поставщика из учёта (только admin)
    if data.startswith("stock_delete_"):
        if role != "admin":
            answer_callback(callback["id"], "⛔ Только Администратор")
            return
        supplier = data[len("stock_delete_"):]
        delete_stock_supplier(chat_id, supplier)
        answer_callback(callback["id"], "🗑 Удалено")
        delete_message(chat_id, message_id)
        return

    answer_callback(callback["id"])


# ===================================================================
#  Управление пользователями по @username
# ===================================================================

def process_username_input(chat_id, user_id, text, state, role):
    clear_user_state(user_id)
    uname = text.strip().lstrip("@").lower()

    if not uname:
        send_message(chat_id, "❌ Введите @username.", reply_markup=keyboard_for_role(role))
        return

    action = state.replace("waiting_username_", "")

    if action == "add_manager":
        add_user_by_username(chat_id, uname, "manager", role)
    elif action == "remove_manager":
        remove_user_by_username(chat_id, uname, "manager", role)
    elif action == "add_storekeeper":
        add_user_by_username(chat_id, uname, "storekeeper", role)
    elif action == "remove_storekeeper":
        remove_user_by_username(chat_id, uname, "storekeeper", role)
    elif action == "delete_user":
        delete_user_by_username(chat_id, user_id, uname, role)


def delete_user_by_username(chat_id, admin_user_id, username, role):
    """Удалить пользователя из системы (только admin)."""
    kb = keyboard_for_role(role)
    with get_session() as session:
        user = session.query(User).filter(User.username == username).first()
        if not user:
            send_message(chat_id, f"❌ Пользователь @{username} не найден.", reply_markup=kb)
            return
        # Защита от самоудаления
        if user.telegram_id == admin_user_id:
            send_message(chat_id, "⚠️ Вы не можете удалить самого себя!", reply_markup=kb)
            return
        uname = user.username
        urole = user.role
        session.delete(user)
    send_message(chat_id, f"✅ Пользователь @{uname} ({urole}) удалён из системы.", reply_markup=kb)


def add_user_by_username(chat_id, username, target_role, caller_role):
    kb = keyboard_for_role(caller_role)
    with get_session() as session:
        exists = session.query(User).filter(User.username == username).first()
        if exists:
            send_message(chat_id, f"ℹ️ @{username} уже в системе (роль: {exists.role}).", reply_markup=kb)
            return
        session.add(User(username=username, role=target_role))
    role_name = "Менеджер" if target_role == "manager" else "Кладовщик"
    send_message(chat_id, f"✅ {role_name} <b>@{username}</b> добавлен.", reply_markup=kb)


def remove_user_by_username(chat_id, username, target_role, caller_role):
    kb = keyboard_for_role(caller_role)
    with get_session() as session:
        user = session.query(User).filter(
            User.username == username, User.role == target_role
        ).first()
        if not user:
            send_message(chat_id, f"❌ @{username} не найден среди {target_role}.", reply_markup=kb)
            return
        session.delete(user)
    role_name = "Менеджер" if target_role == "manager" else "Кладовщик"
    send_message(chat_id, f"✅ {role_name} <b>@{username}</b> удалён.", reply_markup=kb)


def handle_list_users(chat_id):
    with get_session() as session:
        users = session.query(User).order_by(User.role, User.username).all()
        if not users:
            send_message(chat_id, "📋 Список пуст.", reply_markup=keyboard_for_role("admin"))
            return
        lines = []
        for u in users:
            role_emoji = {"manager": "👔", "storekeeper": "👷"}.get(u.role, "❓")
            tg = f" (ID: {u.telegram_id})" if u.telegram_id else ""
            lines.append(f"{role_emoji} @{u.username} — {u.role}{tg}")
        send_message(chat_id, "📋 <b>Пользователи:</b>\n\n" + "\n".join(lines),
                     reply_markup=keyboard_for_role("admin"))


# ===================================================================
#  Создание поставки через Excel
# ===================================================================

def process_excel_upload(chat_id, user_id, doc, role):
    import openpyxl

    clear_user_state(user_id)
    kb = keyboard_for_role(role)

    file_name = doc.get("file_name", "noname.xlsx")
    if not file_name.lower().endswith(".xlsx"):
        send_message(chat_id, "❌ Нужен файл формата .xlsx", reply_markup=kb)
        return

    # Скачать файл
    file_id = doc["file_id"]
    file_data = download_tg_file(file_id)
    if not file_data:
        send_message(chat_id, "❌ Не удалось скачать файл.", reply_markup=kb)
        return

    # Парсинг Excel
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True)
        ws = wb.active
        items = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if len(row) < 2:
                continue
            barcode_val, qty_val = row[0], row[1]
            if barcode_val is None or qty_val is None:
                continue
            barcode = str(int(barcode_val)) if isinstance(barcode_val, float) else str(barcode_val)
            barcode = barcode.strip()
            try:
                qty = int(qty_val)
            except (ValueError, TypeError):
                continue
            # Необязательные столбцы: артикул (3-й) и размер (4-й)
            article_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
            size_val = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
            if qty > 0 and barcode:
                items.append((barcode, qty, article_val, size_val))
        wb.close()
    except Exception as e:
        log(f"Excel parse error: {e}")
        send_message(chat_id, f"❌ Ошибка чтения Excel: {e}", reply_markup=kb)
        return

    if not items:
        send_message(chat_id, "❌ В файле не найдено ни одного товара.\n"
                     "Проверьте формат: 1-й столбец — штрихкод, 2-й — количество.\n"
                     "3-й (артикул) и 4-й (размер) — необязательные.",
                     reply_markup=kb)
        return

    # Группировка дублей баркодов: суммируем количество
    grouped = {}
    for barcode, qty, article_val, size_val in items:
        if barcode in grouped:
            grouped[barcode] = (grouped[barcode][0] + qty, grouped[barcode][1], grouped[barcode][2])
        else:
            grouped[barcode] = (qty, article_val, size_val)
    items = [(bc, info[0], info[1], info[2]) for bc, info in grouped.items()]

    # Название поставки = имя файла без расширения
    shipment_name = os.path.splitext(file_name)[0]

    with get_session() as session:
        # Найти user_id в таблице users
        user = session.query(User).filter(User.telegram_id == user_id).first()
        creator_id = user.id if user else None

        shipment = Shipment(name=shipment_name, status="active", created_by=creator_id)
        session.add(shipment)
        session.flush()

        for barcode, qty, article, size in items:
            session.add(Item(
                shipment_id=shipment.id, barcode=barcode, quantity=qty,
                article=article, size=size,
            ))
        sid = shipment.id

    total_items = sum(q for _, q, _, _ in items)
    send_message(
        chat_id,
        f"✅ Поставка <b>#{sid}</b> создана!\n\n"
        f"📦 Название: <b>{shipment_name}</b>\n"
        f"📋 Артикулов: {len(items)}\n"
        f"📊 Всего единиц: {total_items}",
        reply_markup=kb,
    )


def process_reception_excel(chat_id, user_id, doc, role):
    """Импорт приёмки из Excel: парсинг → Shipment(reception) + ReceptionItems + Stock."""
    import openpyxl

    clear_user_state(user_id)
    kb = keyboard_for_role(role)

    file_name = doc.get("file_name", "noname.xlsx")
    if not file_name.lower().endswith(".xlsx"):
        send_message(chat_id, "❌ Нужен файл формата .xlsx", reply_markup=kb)
        return

    file_id = doc["file_id"]
    file_data = download_tg_file(file_id)
    if not file_data:
        send_message(chat_id, "❌ Не удалось скачать файл.", reply_markup=kb)
        return

    # Парсинг Excel
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True)
        ws = wb.active
        items = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if len(row) < 2:
                continue
            barcode_val, qty_val = row[0], row[1]
            if barcode_val is None or qty_val is None:
                continue
            barcode = str(int(barcode_val)) if isinstance(barcode_val, float) else str(barcode_val)
            barcode = barcode.strip()
            try:
                qty = int(qty_val)
            except (ValueError, TypeError):
                continue
            article_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
            size_val = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
            if qty > 0 and barcode:
                items.append((barcode, qty, article_val, size_val))
        wb.close()
    except Exception as e:
        log(f"Reception Excel parse error: {e}")
        send_message(chat_id, f"❌ Ошибка чтения Excel: {e}", reply_markup=kb)
        return

    if not items:
        send_message(chat_id, "❌ В файле не найдено ни одного товара.", reply_markup=kb)
        return

    # Группировка дублей
    grouped = {}
    for barcode, qty, article_val, size_val in items:
        if barcode in grouped:
            grouped[barcode] = (grouped[barcode][0] + qty, grouped[barcode][1], grouped[barcode][2])
        else:
            grouped[barcode] = (qty, article_val, size_val)
    items = [(bc, info[0], info[1], info[2]) for bc, info in grouped.items()]

    reception_name = os.path.splitext(file_name)[0]

    with get_session() as session:
        user = session.query(User).filter(User.telegram_id == user_id).first()
        creator_id = user.id if user else None

        rec = Shipment(name=reception_name, type="reception", status="archived", created_by=creator_id)
        session.add(rec)
        session.flush()
        rec_id = rec.id

        for barcode, qty, article, size in items:
            session.add(ReceptionItem(
                shipment_id=rec_id, barcode=barcode, quantity=qty,
                article=article, size=size,
            ))
            # Обновить Stock
            stock = session.query(Stock).filter(
                Stock.supplier_name == reception_name,
                Stock.barcode == barcode,
                Stock.is_archived == 0,
            ).first()
            if stock:
                stock.quantity += qty
            else:
                session.add(Stock(
                    supplier_name=reception_name, barcode=barcode,
                    article=article, size=size,
                    quantity=qty,
                ))
            # Движение
            session.add(StockMovement(
                shipment_id=rec_id, movement_type="reception",
                supplier_name=reception_name, barcode=barcode,
                article=article, size=size,
                quantity=qty,
            ))

    total_items = sum(q for _, q, _, _ in items)
    send_message(
        chat_id,
        f"✅ Приёмка <b>#{rec_id}</b> импортирована и заархивирована!\n\n"
        f"📥 Поставщик: <b>{reception_name}</b>\n"
        f"📋 Позиций: {len(items)}\n"
        f"📊 Всего единиц: {total_items}\n\n"
        f"Остатки обновлены в 📊 Учет товара.",
        reply_markup=kb,
    )


# ===================================================================
#  Управление поставками
# ===================================================================

def handle_my_shipments(chat_id, caller_role="manager"):
    """Показать поставки со статусом active и completed."""
    with get_session() as session:
        shipments = (
            session.query(Shipment)
            .filter(Shipment.status.in_(["active", "completed"]), Shipment.type == "shipment")
            .order_by(Shipment.created_at.desc())
            .all()
        )
        if not shipments:
            send_message(chat_id, "📋 Здесь пока ничего нет.")
            return

        for s in shipments:
            total_plan = session.query(func.sum(Item.quantity)).filter(
                Item.shipment_id == s.id).scalar() or 0
            total_packed = session.query(func.count(Packed.id)).filter(
                Packed.shipment_id == s.id).scalar() or 0
            
            # Count actual boxes with items (ignore empty/current if empty)
            box_count = session.query(func.count(distinct(Packed.box_number))).filter(
                Packed.shipment_id == s.id).scalar() or 0

            dt = s.created_at.strftime("%d.%m.%Y") if s.created_at else "—"
            badge = "✅" if s.status == "completed" else "🟢"

            text = (
                f"{badge} <b>#{s.id} — {s.name}</b>  ({dt})\n"
                f"📊 Собрано: {total_packed} / {total_plan}\n"
                f"📦 Коробов: {box_count}"
            )

            # Кнопки: Отчёт | В архив | Удалить(admin)
            buttons_row = [
                {"text": "📊 Отчёт", "callback_data": f"report_shipment_{s.id}"},
                {"text": "📂 В архив", "callback_data": f"archive_shipment_{s.id}"},
            ]
            if caller_role == "admin":
                buttons_row.append(
                    {"text": "🗑 Удалить", "callback_data": f"delete_shipment_{s.id}"},
                )

            send_message(chat_id, text, reply_markup={"inline_keyboard": [buttons_row]})



def handle_my_receptions(chat_id, caller_role="manager"):
    """Показать приёмки со статусом active и completed."""
    with get_session() as session:
        receptions = (
            session.query(Shipment)
            .filter(Shipment.status.in_(["active", "completed"]), Shipment.type == "reception")
            .order_by(Shipment.created_at.desc())
            .all()
        )
        if not receptions:
            send_message(chat_id, "📥 Здесь пока ничего нет.")
            return

        for s in receptions:
            total_scanned = session.query(func.sum(ReceptionItem.quantity)).filter(
                ReceptionItem.shipment_id == s.id).scalar() or 0
            unique_barcodes = session.query(func.count(ReceptionItem.id)).filter(
                ReceptionItem.shipment_id == s.id).scalar() or 0
            dt = s.created_at.strftime("%d.%m.%Y") if s.created_at else "—"
            badge = "✅" if s.status == "completed" else "🟢"

            text = (
                f"{badge} <b>#{s.id} — {s.name}</b>  ({dt})\n"
                f"📊 Всего единиц: <b>{total_scanned}</b>"
            )

            buttons_row = [
                {"text": "📊 Отчёт", "callback_data": f"report_reception_{s.id}"},
                {"text": "📂 В архив", "callback_data": f"archive_reception_{s.id}"},
            ]
            if caller_role == "admin":
                buttons_row.append(
                    {"text": "🗑 Удалить", "callback_data": f"delete_reception_{s.id}"},
                )

            send_message(chat_id, text, reply_markup={"inline_keyboard": [buttons_row]})



def send_shipment_report(chat_id, shipment_id):
    """Отправить Excel-отчёт. Недособранные артикулы подсвечиваются красным."""
    import openpyxl

    with get_session() as session:
        shipment = session.query(Shipment).filter(Shipment.id == shipment_id).first()
        if not shipment:
            send_message(chat_id, "❌ Поставка не найдена.")
            return

        shipment_name = shipment.name

        # План: barcode → {quantity, article, size}
        plan_items = session.query(Item).filter(Item.shipment_id == shipment_id).all()
        plan_map = {}
        for item in plan_items:
            plan_map[item.barcode] = {
                "plan": item.quantity,
                "article": item.article or "—",
                "size": item.size or "—",
            }

        # Факт: barcode → суммарно собрано
        packed_total = {}
        packed_rows_raw = (
            session.query(
                Packed.barcode,
                Packed.box_number,
                func.count(Packed.id).label("qty"),
            )
            .filter(Packed.shipment_id == shipment_id)
            .group_by(Packed.barcode, Packed.box_number)
            .order_by(Packed.box_number)
            .all()
        )
        packed_rows = []
        for r in packed_rows_raw:
            packed_rows.append({"barcode": r.barcode, "qty": r.qty, "box": r.box_number})
            packed_total[r.barcode] = packed_total.get(r.barcode, 0) + r.qty

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Отчёт #{shipment_id}"

    red_fill = openpyxl.styles.PatternFill(start_color="FFCCCC", fill_type="solid")
    header_fill = openpyxl.styles.PatternFill(start_color="D9E2F3", fill_type="solid")
    bold_font = openpyxl.styles.Font(bold=True)

    headers = ["Штрихкод", "План", "Факт", "Короб №", "Артикул", "Размер"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.fill = header_fill

    row_idx = 2
    for pr in packed_rows:
        info = plan_map.get(pr["barcode"], {"plan": 0, "article": "—", "size": "—"})
        total_packed_for_bc = packed_total.get(pr["barcode"], 0)
        is_shortage = total_packed_for_bc < info["plan"]

        ws.cell(row=row_idx, column=1, value=pr["barcode"])
        ws.cell(row=row_idx, column=2, value=info["plan"])
        ws.cell(row=row_idx, column=3, value=pr["qty"])
        ws.cell(row=row_idx, column=4, value=pr["box"])
        ws.cell(row=row_idx, column=5, value=info["article"])
        ws.cell(row=row_idx, column=6, value=info["size"])

        # Подсветка красным если недособрано
        if is_shortage:
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = red_fill

        row_idx += 1

    # Добавить строки для баркодов, которые вообще не сканировали (0 собрано)
    for barcode, info in plan_map.items():
        if barcode not in packed_total:
            ws.cell(row=row_idx, column=1, value=barcode)
            ws.cell(row=row_idx, column=2, value=info["plan"])
            ws.cell(row=row_idx, column=3, value=0)
            ws.cell(row=row_idx, column=4, value="—")
            ws.cell(row=row_idx, column=5, value=info["article"])
            ws.cell(row=row_idx, column=6, value=info["size"])
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = red_fill
            row_idx += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)

    filename = f"{shipment_name}_report_{datetime.now().strftime('%H%M%S')}.xlsx"
    send_document(chat_id, buf.getvalue(), filename, caption=f"📊 Отчёт: {shipment_name}")


def send_reception_report(chat_id, shipment_id):
    """Отправить Excel-отчёт по приёмке (4 колонки)."""
    import openpyxl

    with get_session() as session:
        shipment = session.query(Shipment).filter(Shipment.id == shipment_id).first()
        if not shipment:
            send_message(chat_id, "❌ Приёмка не найдена.")
            return

        shipment_name = shipment.name
        items = (
            session.query(ReceptionItem)
            .filter(ReceptionItem.shipment_id == shipment_id)
            .order_by(ReceptionItem.scanned_at)
            .all()
        )
        rows = [{"barcode": i.barcode, "qty": i.quantity,
                 "article": i.article or "—", "size": i.size or "—"} for i in items]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Приёмка #{shipment_id}"

    header_fill = openpyxl.styles.PatternFill(start_color="D9E2F3", fill_type="solid")
    bold_font = openpyxl.styles.Font(bold=True)

    headers = ["Штрихкод", "Кол-во", "Артикул", "Размер"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.fill = header_fill

    for idx, r in enumerate(rows, 2):
        ws.cell(row=idx, column=1, value=r["barcode"])
        ws.cell(row=idx, column=2, value=r["qty"])
        ws.cell(row=idx, column=3, value=r["article"])
        ws.cell(row=idx, column=4, value=r["size"])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)

    filename = f"{shipment_name}_reception_{datetime.now().strftime('%H%M%S')}.xlsx"
    send_document(chat_id, buf.getvalue(), filename, caption=f"📊 Приёмка: {shipment_name}")


def delete_shipment(chat_id, shipment_id):
    """Полное удаление поставки (Admin only — проверка в callback)."""
    with get_session() as session:
        shipment = session.query(Shipment).filter(Shipment.id == shipment_id).first()
        if not shipment:
            send_message(chat_id, "❌ Поставка не найдена.")
            return
        name = shipment.name
        session.delete(shipment)
    send_message(chat_id, f"🗑 Поставка <b>#{shipment_id} — {name}</b> полностью удалена.")


def archive_shipment(chat_id, shipment_id):
    """completed → archived + обновление Stock."""
    with get_session() as session:
        shipment = session.query(Shipment).filter(Shipment.id == shipment_id).first()
        if not shipment:
            send_message(chat_id, "❌ Поставка не найдена.")
            return
        shipment.status = "archived"
        name = shipment.name
        stype = shipment.type or "shipment"

        # --- Обновление склада ---
        if stype == "reception":
            # Приёмка: +кол-во в Stock
            for ri in shipment.reception_items:
                stock = session.query(Stock).filter(
                    Stock.supplier_name == name, Stock.barcode == ri.barcode
                ).first()
                if stock:
                    stock.quantity += ri.quantity
                else:
                    session.add(Stock(
                        supplier_name=name, barcode=ri.barcode,
                        article=ri.article, size=ri.size,
                        quantity=ri.quantity,
                    ))
                session.add(StockMovement(
                    shipment_id=shipment_id, movement_type="reception",
                    supplier_name=name, barcode=ri.barcode,
                    article=ri.article, size=ri.size,
                    quantity=ri.quantity,
                ))
        elif stype == "shipment":
            # Отгрузка: -кол-во из Stock (по факту упаковки)
            packed_agg = (
                session.query(Packed.barcode, func.count(Packed.id).label("qty"))
                .filter(Packed.shipment_id == shipment_id)
                .group_by(Packed.barcode)
                .all()
            )
            for barcode, qty in packed_agg:
                # Найти артикул/размер из плана
                plan_item = session.query(Item).filter(
                    Item.shipment_id == shipment_id, Item.barcode == barcode
                ).first()
                article = plan_item.article if plan_item else None
                size_val = plan_item.size if plan_item else None

                stocks = session.query(Stock).filter(Stock.barcode == barcode).all()
                remaining = qty
                for s in stocks:
                    if remaining <= 0:
                        break
                    deduct = min(s.quantity, remaining)
                    s.quantity -= deduct
                    remaining -= deduct

                session.add(StockMovement(
                    shipment_id=shipment_id, movement_type="shipment",
                    supplier_name=name, barcode=barcode,
                    article=article, size=size_val,
                    quantity=-qty,
                ))

    send_message(chat_id, f"📂 <b>#{shipment_id} — {name}</b> перемещено в историю.")


def reopen_shipment(chat_id, shipment_id):
    """archived → completed (Admin only — проверка в callback)."""
    with get_session() as session:
        shipment = session.query(Shipment).filter(Shipment.id == shipment_id).first()
        if not shipment:
            send_message(chat_id, "❌ Поставка не найдена.")
            return
        if shipment.status != "archived":
            send_message(chat_id, "ℹ️ Поставка не в архиве.")
            return
        shipment.status = "completed"
        name = shipment.name
    send_message(chat_id, f"🔄 Поставка <b>#{shipment_id} — {name}</b> возвращена в Мои поставки.")


def handle_archive(chat_id, caller_role="manager"):
    """Показать поставки и приёмки со статусом archived."""
    try:
        data_rows = []
        with get_session() as session:
            shipments = (
                session.query(Shipment)
                .filter(Shipment.status == "archived")
                .order_by(Shipment.created_at.desc())
                .all()
            )
            for s in shipments:
                data_rows.append({
                    "id": s.id,
                    "name": s.name,
                    "type": s.type or "shipment",
                    "dt": s.created_at.strftime("%d.%m.%Y") if s.created_at else "—",
                })

        if not data_rows:
            send_message(chat_id, "📂 Здесь пока ничего нет.")
            return

        lines = ["📂 <b>Архив:</b>\n"]
        buttons = []
        for d in data_rows:
            icon = "📦" if d["type"] == "shipment" else "📥"
            lines.append(f"  {icon} <b>#{d['id']}</b> — {d['name']}  ({d['dt']})")
            report_cb = f"report_{d['type']}_{d['id']}"
            row = [{"text": "📊 Отчёт", "callback_data": report_cb}]
            if caller_role == "admin":
                row.append({"text": "🔄 Вернуть", "callback_data": f"reopen_{d['id']}"})
            buttons.append(row)

        send_message(chat_id, "\n".join(lines), reply_markup={"inline_keyboard": buttons})
    except Exception as e:
        log(f"handle_archive error: {e}")
        send_message(chat_id, "❌ Ошибка при загрузке архива.")


def handle_history_shipments(chat_id, caller_role="manager"):
    """Показать архив поставок."""
    with get_session() as session:
        shipments = (
            session.query(Shipment)
            .filter(Shipment.status == "archived", Shipment.type == "shipment")
            .order_by(Shipment.created_at.desc())
            .limit(20)
            .all()
        )
        if not shipments:
            send_message(chat_id, "� Архив поставок пуст.")
            return

        lines = ["� <b>История поставок:</b>\n"]
        buttons = []
        for s in shipments:
            dt = s.created_at.strftime("%d.%m.%Y") if s.created_at else "—"
            lines.append(f"📦 <b>#{s.id} — {s.name}</b> ({dt})")
            row = [{"text": "� Скачать отчёт", "callback_data": f"report_shipment_{s.id}"}]
            if caller_role == "admin":
                 row.append({"text": "� Вернуть", "callback_data": f"reopen_{s.id}"})
            buttons.append(row)
        
        send_message(chat_id, "\n".join(lines), reply_markup={"inline_keyboard": buttons})


def handle_history_receptions(chat_id, caller_role="manager"):
    """Показать архив приёмок."""
    with get_session() as session:
        receptions = (
            session.query(Shipment)
            .filter(Shipment.status == "archived", Shipment.type == "reception")
            .order_by(Shipment.created_at.desc())
            .limit(20)
            .all()
        )
        if not receptions:
            send_message(chat_id, "📜 Архив приёмок пуст.")
            return

        lines = ["� <b>История приёмок:</b>\n"]
        buttons = []
        for s in receptions:
            dt = s.created_at.strftime("%d.%m.%Y") if s.created_at else "—"
            lines.append(f"📥 <b>#{s.id} — {s.name}</b> ({dt})")
            row = [{"text": "� Скачать отчёт", "callback_data": f"report_reception_{s.id}"}]
            if caller_role == "admin":
                 row.append({"text": "� Вернуть", "callback_data": f"reopen_{s.id}"})
            buttons.append(row)

        send_message(chat_id, "\n".join(lines), reply_markup={"inline_keyboard": buttons})


# ===================================================================
#  Складской учёт
# ===================================================================

def handle_stock(chat_id, caller_role="manager"):
    """📊 Учет товара: список активных поставщиков с кнопками отчёта/архива."""
    with get_session() as session:
        suppliers = (
            session.query(
                Stock.supplier_name,
                func.sum(Stock.quantity).label("total"),
                func.max(Stock.updated_at).label("last_upd")
            )
            .filter(Stock.is_archived == 0)
            .group_by(Stock.supplier_name)
            .order_by(Stock.supplier_name)
            .all()
        )
    if not suppliers:
        send_message(chat_id, "📊 Склад пуст. Данные появятся после первой архивации приёмки.")
        return

    lines = ["� <b>Учёт товара по поставщикам:</b>\n"]
    buttons = []
    for name, total, last_upd in suppliers:
        dt = last_upd.strftime("%d.%m.%Y %H:%M") if last_upd else "—"
        lines.append(f"  📦 <b>{name}</b> — {total} ед. (Обн: {dt})")
        row = [
            {"text": "📊 Отчёт", "callback_data": f"stock_report_{name}"},
            {"text": "📂 В архив", "callback_data": f"stock_archive_{name}"},
        ]
        if caller_role == "admin":
            row.append({"text": "� Удалить", "callback_data": f"stock_delete_{name}"})
        buttons.append(row)

    send_message(chat_id, "\n".join(lines), reply_markup={"inline_keyboard": buttons})


def handle_stock_history(chat_id, caller_role="manager"):
    """📜 История: выбор раздела."""
    buttons = [
        [{"text": "📜 История поставок", "callback_data": "history_shipments"}],
        [{"text": "📜 История приёмок", "callback_data": "history_receptions"}],
    ]
    send_message(chat_id, "📜 Выберите раздел истории:", reply_markup={"inline_keyboard": buttons})


def archive_stock_supplier(chat_id, supplier_name):
    """Переместить поставщика в архив склада."""
    with get_session() as session:
        updated = (
            session.query(Stock)
            .filter(Stock.supplier_name == supplier_name, Stock.is_archived == 0)
            .update({Stock.is_archived: 1})
        )
    if updated:
        send_message(chat_id, f"📂 Поставщик <b>{supplier_name}</b> перемещён в архив.")
    else:
        send_message(chat_id, "❌ Поставщик не найден.")


def unarchive_stock_supplier(chat_id, supplier_name):
    """Вернуть поставщика из архива."""
    with get_session() as session:
        updated = (
            session.query(Stock)
            .filter(Stock.supplier_name == supplier_name, Stock.is_archived == 1)
            .update({Stock.is_archived: 0})
        )
    if updated:
        send_message(chat_id, f"🔄 Поставщик <b>{supplier_name}</b> возвращён в учёт.")
    else:
        send_message(chat_id, "❌ Поставщик не найден.")


def delete_stock_supplier(chat_id, supplier_name):
    """Полностью удалить записи поставщика (только admin)."""
    with get_session() as session:
        deleted = (
            session.query(Stock)
            .filter(Stock.supplier_name == supplier_name)
            .delete()
        )
    if deleted:
        send_message(chat_id, f"🗑 Поставщик <b>{supplier_name}</b> полностью удалён.")
    else:
        send_message(chat_id, "❌ Поставщик не найден.")


def send_stock_report(chat_id, supplier_name):
    """Отправить Excel-отчёт по остаткам поставщика."""
    import openpyxl

    with get_session() as session:
        items = (
            session.query(Stock)
            .filter(Stock.supplier_name == supplier_name)
            .order_by(Stock.barcode)
            .all()
        )
        rows = [{"barcode": s.barcode, "article": s.article or "—",
                 "size": s.size or "—", "qty": s.quantity} for s in items]

    if not rows:
        send_message(chat_id, f"📊 У поставщика «{supplier_name}» нет записей.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Остатки — {supplier_name}"

    header_fill = openpyxl.styles.PatternFill(start_color="D9E2F3", fill_type="solid")
    bold_font = openpyxl.styles.Font(bold=True)

    headers = ["Штрихкод", "Артикул", "Размер", "Остаток"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.fill = header_fill

    for idx, r in enumerate(rows, 2):
        ws.cell(row=idx, column=1, value=r["barcode"])
        ws.cell(row=idx, column=2, value=r["article"])
        ws.cell(row=idx, column=3, value=r["size"])
        ws.cell(row=idx, column=4, value=r["qty"])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)

    filename = f"stock_{supplier_name}_{datetime.now().strftime('%H%M%S')}.xlsx"
    send_document(chat_id, buf.getvalue(), filename, caption=f"📊 Остатки: {supplier_name}")


def handle_history_shipments(chat_id, caller_role="manager"):
    """Показать архив поставок."""
    with get_session() as session:
        shipments = (
            session.query(Shipment)
            .filter(Shipment.status == "archived", Shipment.type == "shipment")
            .order_by(Shipment.created_at.desc())
            .limit(20)
            .all()
        )
        if not shipments:
            send_message(chat_id, "📜 Архив поставок пуст.")
            return

        lines = ["📜 <b>История поставок:</b>\n"]
        buttons = []
        for s in shipments:
            dt = s.created_at.strftime("%d.%m.%Y") if s.created_at else "—"
            lines.append(f"📦 <b>#{s.id} — {s.name}</b> ({dt})")
            row = [{"text": "📥 Скачать отчёт", "callback_data": f"report_shipment_{s.id}"}]
            if caller_role == "admin":
                 row.append({"text": "🔄 Вернуть", "callback_data": f"reopen_{s.id}"})
            buttons.append(row)
        
        send_message(chat_id, "\n".join(lines), reply_markup={"inline_keyboard": buttons})


def handle_history_receptions(chat_id, caller_role="manager"):
    """Показать архив приёмок."""
    with get_session() as session:
        receptions = (
            session.query(Shipment)
            .filter(Shipment.status == "archived", Shipment.type == "reception")
            .order_by(Shipment.created_at.desc())
            .limit(20)
            .all()
        )
        if not receptions:
            send_message(chat_id, "📜 Архив приёмок пуст.")
            return

        lines = ["📜 <b>История приёмок:</b>\n"]
        buttons = []
        for s in receptions:
            dt = s.created_at.strftime("%d.%m.%Y") if s.created_at else "—"
            lines.append(f"📥 <b>#{s.id} — {s.name}</b> ({dt})")
            row = [{"text": "📥 Скачать отчёт", "callback_data": f"report_reception_{s.id}"}]
            if caller_role == "admin":
                 row.append({"text": "🔄 Вернуть", "callback_data": f"reopen_{s.id}"})
            buttons.append(row)

        send_message(chat_id, "\n".join(lines), reply_markup={"inline_keyboard": buttons})


# ===================================================================
#  API для Mini App (JSON)
# ===================================================================

def api_shipments():
    """GET: список активных поставок."""
    with get_session() as session:
        shipments = session.query(Shipment).filter(Shipment.status == "active", Shipment.type == "shipment").all()
        result = [
            {
                "id": s.id,
                "name": s.name,
                "last_box_number": s.last_box_number,
                "created_at": s.created_at.isoformat() if s.created_at else "",
            }
            for s in shipments
        ]
    respond_json(result)


def api_shipment_detail(shipment_id):
    """GET: детали поставки (план + прогресс)."""
    with get_session() as session:
        shipment = session.query(Shipment).filter(Shipment.id == shipment_id).first()
        if not shipment:
            respond_json({"error": "Поставка не найдена"}, "404 Not Found")
            return

        items_data = []
        for item in shipment.items:
            packed_count = (
                session.query(func.count(Packed.id))
                .filter(Packed.shipment_id == shipment_id, Packed.barcode == item.barcode)
                .scalar()
            )
            items_data.append({
                "id": item.id,
                "barcode": item.barcode,
                "quantity": item.quantity,
                "packed": packed_count,
                "remaining": item.quantity - packed_count,
                "article": item.article or "—",
                "size": item.size or "—",
            })

        respond_json({
            "id": shipment.id,
            "name": shipment.name,
            "status": shipment.status,
            "last_box_number": shipment.last_box_number,
            "items": items_data,
        })


def api_scan():
    """POST: обработка скана товара."""
    body = read_request_body()
    try:
        data = json.loads(body)
    except Exception:
        respond_json({"status": "error", "message": "Неверный JSON"}, "400 Bad Request")
        return

    shipment_id = data.get("shipment_id")
    box_number = data.get("box_number")
    item_barcode = (data.get("item_barcode") or "").strip()
    quantity = int(data.get("quantity", 1) or 1)
    if quantity < 1:
        quantity = 1

    if not all([shipment_id, box_number, item_barcode]):
        respond_json({"status": "error", "message": "Недостаточно данных"}, "400 Bad Request")
        return

    with get_session() as session:
        item = (
            session.query(Item)
            .filter(Item.shipment_id == shipment_id, Item.barcode == item_barcode)
            .first()
        )

        if not item:
            respond_json({
                "status": "error",
                "message": f"Артикул {item_barcode} не найден в плане",
            })
            return

        packed_count = (
            session.query(func.count(Packed.id))
            .filter(Packed.shipment_id == shipment_id, Packed.barcode == item_barcode)
            .scalar()
        )

        if packed_count + quantity > item.quantity:
            respond_json({
                "status": "error",
                "message": f"Перебор! {item_barcode}: план {item.quantity}, уже {packed_count}, добавляете {quantity}",
            })
            return

        for _ in range(quantity):
            session.add(Packed(
                shipment_id=shipment_id,
                item_id=item.id,
                barcode=item_barcode,
                box_number=int(box_number),
            ))

        remaining = item.quantity - packed_count - quantity

    respond_json({"status": "ok", "remaining": remaining})


def api_open_box():
    """POST: получить текущий номер короба (или создать первый)."""
    body = read_request_body()
    try:
        data = json.loads(body)
    except Exception:
        respond_json({"status": "error", "message": "Неверный JSON"}, "400 Bad Request")
        return

    shipment_id = data.get("shipment_id")
    if not shipment_id:
        respond_json({"status": "error", "message": "shipment_id обязателен"}, "400 Bad Request")
        return

    with get_session() as session:
        shipment = session.query(Shipment).filter(Shipment.id == shipment_id).first()
        if not shipment:
            respond_json({"status": "error", "message": "Поставка не найдена"}, "404 Not Found")
            return

        if shipment.last_box_number == 0:
            shipment.last_box_number = 1

        respond_json({"status": "ok", "box_number": shipment.last_box_number})


def api_close_box():
    """POST: закрыть текущий короб → следующий номер."""
    body = read_request_body()
    try:
        data = json.loads(body)
    except Exception:
        respond_json({"status": "error", "message": "Неверный JSON"}, "400 Bad Request")
        return

    shipment_id = data.get("shipment_id")
    if not shipment_id:
        respond_json({"status": "error", "message": "shipment_id обязателен"}, "400 Bad Request")
        return

    with get_session() as session:
        shipment = session.query(Shipment).filter(Shipment.id == shipment_id).first()
        if not shipment:
            respond_json({"status": "error", "message": "Поставка не найдена"}, "404 Not Found")
            return

        shipment.last_box_number += 1
        new_box = shipment.last_box_number

        # Проверяем, все ли товары полностью собраны
        all_complete = True
        for item in shipment.items:
            packed_count = (
                session.query(func.count(Packed.id))
                .filter(Packed.shipment_id == shipment_id, Packed.barcode == item.barcode)
                .scalar()
            )
            if packed_count < item.quantity:
                all_complete = False
                break

        if all_complete and len(shipment.items) > 0:
            shipment.status = "completed"

    respond_json({"status": "ok", "box_number": new_box, "completed": all_complete})


# ===================================================================
#  API для Mini App — Приёмка
# ===================================================================

def api_create_reception():
    """POST: создать приёмку (только имя)."""
    body = json.loads(read_request_body())
    name = body.get("name", "").strip()
    if not name:
        respond_json({"status": "error", "message": "Укажите название приёмки"}, "400 Bad Request")
        return
    with get_session() as session:
        rec = Shipment(name=name, type="reception", status="active")
        session.add(rec)
        session.flush()
        rid = rec.id
    respond_json({"status": "ok", "id": rid})


def api_reception_scan():
    """POST: сканирование ШК в приёмку (upsert)."""
    body = json.loads(read_request_body())
    reception_id = body.get("reception_id")
    barcode = str(body.get("barcode", "")).strip()
    if not reception_id or not barcode:
        respond_json({"status": "error", "message": "reception_id и barcode обязательны"}, "400 Bad Request")
        return

    with get_session() as session:
        rec = session.query(Shipment).filter(Shipment.id == reception_id, Shipment.type == "reception").first()
        if not rec:
            respond_json({"status": "error", "message": "Приёмка не найдена"}, "404 Not Found")
            return

        existing = (
            session.query(ReceptionItem)
            .filter(ReceptionItem.shipment_id == reception_id, ReceptionItem.barcode == barcode)
            .first()
        )
        scan_qty = int(body.get("quantity", 1))
        if scan_qty < 1:
            scan_qty = 1

        if existing:
            existing.quantity += scan_qty
            existing.scanned_at = datetime.utcnow()
            qty = existing.quantity
        else:
            item = ReceptionItem(shipment_id=reception_id, barcode=barcode, quantity=scan_qty)
            session.add(item)
            qty = scan_qty

    respond_json({"status": "ok", "barcode": barcode, "quantity": qty})


def api_receptions():
    """GET: список активных приёмок."""
    with get_session() as session:
        receptions = session.query(Shipment).filter(
            Shipment.status == "active", Shipment.type == "reception"
        ).all()
        result = [
            {"id": r.id, "name": r.name, "created_at": r.created_at.isoformat() if r.created_at else ""}
            for r in receptions
        ]
    respond_json(result)


def api_reception_detail(reception_id):
    """GET: детали приёмки (список отсканированных ШК)."""
    with get_session() as session:
        rec = session.query(Shipment).filter(Shipment.id == reception_id, Shipment.type == "reception").first()
        if not rec:
            respond_json({"error": "Приёмка не найдена"}, "404 Not Found")
            return

        items = (
            session.query(ReceptionItem)
            .filter(ReceptionItem.shipment_id == reception_id)
            .order_by(ReceptionItem.scanned_at.desc())
            .all()
        )
        items_data = [
            {
                "barcode": i.barcode,
                "quantity": i.quantity,
                "article": i.article or "—",
                "size": i.size or "—",
            }
            for i in items
        ]

    respond_json({
        "id": rec.id,
        "name": rec.name,
        "status": rec.status,
        "items": items_data,
    })


def serve_mini_app():
    """GET: отдать HTML страницу Mini App."""
    template_path = os.path.join(SCRIPT_DIR, "templates", "index.html")
    try:
        with open(template_path, "r", encoding="utf-8") as f:
            html = f.read()
        respond_html(html)
    except FileNotFoundError:
        respond_html("<h1>Ошибка: шаблон index.html не найден</h1>", "500 Internal Server Error")


# ===================================================================
#  LONG POLLING — ТОЧКА ВХОДА
# ===================================================================
import time

def start_polling():
    """Фоновый сбор обновлений Telegram без Webhook (Long Polling)"""
    # Удаляем старый вебхук перед запуском поллинга
    tg_api("deleteWebhook")
    log("Webhook removed, started long polling...")
    print("Бот успешно запущен в режиме Long Polling!")
    
    offset = None
    while True:
        try:
            payload = {"timeout": 30}
            if offset:
                payload["offset"] = offset
            
            updates = tg_api("getUpdates", payload)
            if updates and "result" in updates:
                for update in updates["result"]:
                    offset = update["update_id"] + 1
                    handle_telegram_update(update)
        except Exception as e:
            log(f"Polling error: {e}")
            time.sleep(5)

if __name__ == "__main__":
    start_polling()
